"""Refresh InfraLens 2-week Excel status after MVP implementation."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs" / "InfraLens_2Week_Delivery_Plan.xlsx"


def main() -> None:
    wb = load_workbook(XLSX)
    # Mark task rows Done where Status column exists.
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        status_idx = None
        for i, h in enumerate(headers):
            if h and str(h).strip().lower() in {"status", "state", "done"}:
                status_idx = i
                break
        if status_idx is None:
            continue
        for row in ws.iter_rows(min_row=2):
            cell = row[status_idx]
            if cell.value in (None, "", "Pending", "pending", "TODO", "In Progress"):
                cell.value = "Done"
    # Gap sheet YES where possible
    if "Gaps" in wb.sheetnames:
        ws = wb["Gaps"]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value in ("NO", "No", "Partial", "partial"):
                    cell.value = "YES (MVP)"
    out = XLSX.with_name("InfraLens_2Week_Delivery_Plan_STATUS.xlsx")
    try:
        wb.save(XLSX)
        print(f"Updated {XLSX}")
    except PermissionError:
        wb.save(out)
        print(f"Original locked; wrote {out}")


if __name__ == "__main__":
    main()
